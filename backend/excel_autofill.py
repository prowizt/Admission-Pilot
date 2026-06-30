import io
import json
import re
import os
import time
import base64
import urllib.parse
import openpyxl
import asyncio
from google.api_core.retry_async import AsyncRetry
from decimal import Decimal
from fastapi import APIRouter, UploadFile, File, HTTPException, Header, Form
import google.generativeai as genai

router = APIRouter()

def extract_excel_structure(ws):
    """
    빈 셀의 좌표와 해당 셀이 위치한 행/열의 문맥(헤더) 정보를 추출합니다.
    """
    empty_cells = []
    max_row = min(ws.max_row, 200)
    max_col = min(ws.max_column, 200)
    # 1. 병합된 셀(Merged Cells) 완벽 평탄화 (Flatten)
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col_idx, max_row_idx = merged_range.bounds
        top_left_val = ws.cell(row=min_row, column=min_col).value
        if top_left_val is not None and str(top_left_val).strip() != "":
            # 병합 영역 내의 모든 셀에 좌상단 값 할당 (평탄화)
            for r in range(min_row, max_row_idx + 1):
                for c in range(min_col, max_col_idx + 1):
                    merged_values[(r, c)] = str(top_left_val).strip()

    headers = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in merged_values:
                headers[(r, c)] = merged_values[(r, c)]
            else:
                val = ws.cell(row=r, column=c).value
                if val is not None and str(val).strip() != "":
                    headers[(r, c)] = str(val).strip()

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            
            # 헤더로 판별된 셀은 빈 셀 목록에서 제외
            if (r, c) in headers:
                continue

            # 빈칸(입력칸)이 병합된 경우, 좌상단 1개만 수집하고 나머지 껍데기는 무시
            if type(cell).__name__ == 'MergedCell':
                continue
                
            if cell.value is None or str(cell.value).strip() == "":
                left_headers = []
                for i in range(1, c):
                    if (r, i) in headers:
                        h = headers[(r, i)]
                        # 연속된 동일 헤더 제거 (평탄화로 인한 중복 방지)
                        if not left_headers or left_headers[-1] != h:
                            left_headers.append(h)

                top_headers = []
                for i in range(1, r):
                    if (i, c) in headers:
                        h = headers[(i, c)]
                        if not top_headers or top_headers[-1] != h:
                            top_headers.append(h)
                
                # 좌측 헤더나 상단 헤더 중 하나라도 있으면 유효한 빈 셀로 간주 (유연성 확보)
                if left_headers or top_headers:
                    empty_cells.append({
                        "coord": cell.coordinate,
                        "row_headers": " > ".join(left_headers),
                        "col_headers": " > ".join(top_headers)
                    })
    # 기존에 입력된 데이터(주로 헤더) 정보를 문자열로 요약
    existing_data_summary = []
    for (r, c), val in headers.items():
        existing_data_summary.append(f"셀({ws.cell(row=r, column=c).coordinate}): {val}")
        
    return empty_cells, "\n".join(existing_data_summary)

@router.post("/chat-excel-autofill")
async def chat_excel_autofill(
    file: UploadFile = File(...), 
    question: str = Form(...),
    user_role: str = Form("staff"),
    model_name: str = Form("gemini-2.5-flash"),
    history: str = Form("[]"),
    scraped_context: str = Form(None),
    scraped_file_name: str = Form(None),
    router_model_name: str = Form(None),
    x_gemini_key: str = Header(None)
):
    # 순환 참조 방지를 위해 함수 내부에서 import
    from main import get_mssql_connection, get_dynamic_db_schema, format_sql_query, mask_pii

    start_time = time.time()
    print(f"\n[EXCEL-AUTOFILL] 🚀 하이브리드 채팅 엑셀 자동 채우기 요청 수신 (질문: {question})")
    
    if question:
        question = mask_pii(question)
        
    if not x_gemini_key:
        raise HTTPException(status_code=401, detail="Gemini API Key가 필요합니다.")
        
    try:
        genai.configure(api_key=x_gemini_key)
        conn = get_mssql_connection()
        dynamic_schema = get_dynamic_db_schema(conn, user_role) if conn else "DB 연결 불가"
        
        pre_supplemental_text = ""
        if conn:
            try:
                pre_cursor = conn.cursor()
                pre_cursor.execute("SELECT category, content FROM Sys_SupplementalKnowledge WHERE is_active = 'Y' ORDER BY category ASC, id ASC")
                for r in pre_cursor.fetchall():
                    pre_supplemental_text += f"- [{r[0]}] {r[1]}\n"
            except Exception as e:
                print(f"사전지식 선행 조회 오류: {e}")
        
        # 1. 엑셀 파싱
        file_bytes = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        empty_cells_info, existing_data_context = extract_excel_structure(ws)


        print(f"[EXCEL-AUTOFILL] 📊 1단계: 엑셀 파악 완료 (빈 셀 {len(empty_cells_info)}개, 기존 데이터 {len(existing_data_context.split(chr(10))) if existing_data_context else 0}개 감지)")
        
        # 2. Text-to-SQL (라우터) 구동
        router_model_name = router_model_name or model_name
        print(f"[EXCEL-AUTOFILL] 🧠 2단계: AI 통계 SQL 쿼리 설계 중... (사용 모델: {router_model_name})")
        sql_router_prompt = f"""
        너는 대동대학교 입시처의 [엑셀 매핑 전문 데이터 아키텍트]야.
        사용자가 첨부한 엑셀 파일의 빈 칸을 채우기 위해 필요한 통계 데이터를 DB에서 추출하는 T-SQL 쿼리를 하나 작성해야 해.

        [관리자 등록 최상위 사전지식]
        {pre_supplemental_text if pre_supplemental_text else "등록된 사전지식 없음"}
        - 위 [관리자 등록 최상위 사전지식]에 명시된 규칙은 질문의 의도 파악과 SQL 쿼리 설계 시 절대적으로 준수해야 하는 1순위 예외 규칙입니다.
        
        [가장 중요한 SQL 작성 규칙]
        1. 🚨 **[전처리(Pre-mapping) 장려 및 줄바꿈 절대 금지]** 제공된 [엑셀 빈 셀 구조]의 헤더 명칭(예: '대학졸업자_2년제졸업자')을 참고하여, DB의 원본 데이터를 `CASE WHEN` 등을 사용해 엑셀 키워드와 최대한 똑같이 맞춰서(가공해서) 출력해 주는 것을 적극 권장합니다. **단, 쿼리가 길어지더라도 따옴표('') 내부 문자열에 절대 줄바꿈(Enter/Newline)을 넣지 마십시오!** (문법 에러 발생 원인). 컬럼의 Alias(AS `[전형명]`)는 반드시 `Sys_ColumnCatalog` 스키마에 존재하는 원본 컬럼명만 사용해야 하며, 엑셀 구조를 흉내 내려고 가로로 PIVOT 하지 말고 세로(Row) 방향으로 추출하세요.
        2. 컬럼명이 한글일 수 있으므로 대괄호 []를 반드시 사용해.
        3. 🚨 **[띄어쓰기 절대 금지]** 대괄호 `[]` 내부에는 **어떤 이유로도 단 한 칸의 공백도 넣지 마세요!** (예: `[ 수험번호]`, `[전형 명]`, `[지원인 원]`처럼 공백이 하나라도 들어가면 치명적 에러가 발생합니다. 무조건 `[수험번호]` 처럼 공백을 완전히 붙이세요.)
        4. **[별칭(Alias) 공백 금지 및 대괄호 필수]** `AS [지원_인원]` 처럼 별칭 대괄호 안에도 띄어쓰기를 절대 금지합니다.
        5. **[문자열 비교 공백 엄수]** 조건절에서 문자열을 비교할 때 카탈로그에 정의된 문자열 구조를 임의로 띄어쓰지 마세요. 무조건 원본과 똑같이 붙여 쓰세요.
        6. **[신입/편입 구분 및 정원내/외 조건]** 통계를 낼 때 신입생과 편입생, 정원내/외를 명확하게 분리하세요.
        7. **[LIKE 검색어 공백 절대 금지]** `LIKE` 검색 시 `LIKE '%수능 전형%'` 처럼 단어 사이에 공백을 임의로 넣지 마세요.
        8. **[학년도 자동 추출 강제]** 사용자가 명시적으로 학년도를 말하지 않더라도, 파일명이나 질문의 연도를 찾아 `WHERE [입시학년도] = '추출된연도'` 과 같이 조건을 넣으세요.
        9. 🚨 **[UNION ALL 더미 행 생성 절대 금지]** 엑셀의 빈 셀 구조를 똑같이 따라 하겠다고 0으로 채워진 야간 학과나 더미 행을 `UNION ALL`로 억지로 생성하지 마십시오! DB에 실재하는 학과의 데이터만 단순 `GROUP BY`로 뽑아주면 빈 셀(0값) 매핑은 파이썬이 알아서 완벽히 처리합니다. 
        10. **[ORDER BY 절대 사용 금지]** MS-SQL 오류 방지를 위해 쿼리 마지막에 `ORDER BY` 절을 절대로 작성하지 마십시오!
        11. 🚨 **[일대다 조인 뻥튀기 방지 및 원본 컬럼명 보존]** 기준 정보 테이블과 상세 내역 테이블을 JOIN할 때 1:N 관계로 인해 숫자가 중복(뻥튀기)되지 않도록 주의하십시오. 이를 방지하기 위해 그룹핑 과정에서 가상의 세부 명칭(예: `[세부분류]`)을 생성하여 쪼갰다면, 최종 SELECT 절에서 임의로 `[출력명칭]` 같은 가짜 별칭을 만들지 마십시오. 반드시 매핑 시스템이 인식할 수 있도록 **원래 스키마에 존재하던 원본 컬럼명(Alias)**을 그대로 사용하여 덮어씌워야 합니다. 가짜 컬럼명을 반환하면 매핑 시스템이 데이터를 찾지 못해 모두 0으로 처리됩니다. (올바른 예: `COALESCE([A].[세부분류], [Q].[기준컬럼]) AS [기준컬럼]`)
        12. 🚨 **[절대 PIVOT 금지 및 스키마 원본 컬럼 강제]** 제공된 [키워드]들을 토대로 쿼리의 SELECT 컬럼(Alias)으로 억지로 PIVOT을 만들면 파이썬 서버에서 즉각 500 에러가 터집니다! 가로로 펼치는(PIVOT) 쿼리나 스키마에 없는 이상한 파생 컬럼을 억지로 생성하는 것을 엄격히 금지합니다. 반드시 **Sys_ColumnCatalog에 정의된 원본 컬럼명**만을 사용하여 세로 방향(Row 단위)의 정규화된(Normalized) 기본 1차원 테이블 구조로만 결과를 출력해야 합니다.
        13. 🚨 **[GROUP BY 8120 오류 완벽 방지 (반복 금지)]** MS-SQL에서 `SELECT` 절에 긴 `CASE WHEN` 구문을 사용하고 `GROUP BY` 절에 그대로 복사해서 넣을 경우, 줄바꿈/띄어쓰기 불일치로 8120 에러가 폭발적으로 발생합니다. 이 치명적 오류를 원천 차단하기 위해, **`GROUP BY` 절 내부에 `CASE WHEN` 구문을 사용하는 것을 엄격히 금지**합니다! 반드시 CTE나 인라인 뷰(서브쿼리)를 사용하여 `CASE WHEN`으로 가공된 컬럼(예: `[가공된_컬럼명]`)을 먼저 정의한 뒤, 메인 쿼리에서는 단순히 `GROUP BY [기준컬럼1], [가공된_컬럼명]` 형태로 깔끔하게 묶으십시오.
        14. 🚨 **[문자열 숫자형 변환 필수 (SUM 에러 방지)]** 스키마에서 수치 데이터를 나타내는 컬럼이 `nvarchar` 타입일 경우, `SUM()` 집계 함수를 사용할 때 원본 그대로 쓰지 마십시오! 반드시 `SUM(CASE WHEN ISNUMERIC([문자형수치컬럼])=1 THEN CAST([문자형수치컬럼] AS INT) ELSE 0 END)` 형태로 안전하게 형변환(Casting)하여 8117 에러를 완벽히 차단하십시오.
        15. 🚨 **[합계(Total) 자동 산출 (ROLLUP 필수)]** [엑셀 빈 셀 구조]를 살펴봤을 때 '합계', '계', '소계', '전체' 등의 집계(Total) 칸이 존재한다면, 파이썬 단의 매핑 AI가 직접 계산하지 않도록 **쿼리의 `GROUP BY` 절에 반드시 `WITH ROLLUP` (또는 CUBE) 구문을 추가**하여 DB 단에서 완벽하게 계산된 합계 행(Row)을 함께 추출해 주십시오.

        [데이터베이스 스키마 및 힌트]
        (아래 스키마의 ai_description에 적힌 값(힌트)들을 꼼꼼히 확인하고 쿼리를 작성하세요)
        {dynamic_schema}

        [첨부 엑셀 파일명 (기준 학년도/학기 유추에 활용)]
        {file.filename}

        [엑셀 빈 셀 구조 (이 셀들을 채우기 위한 데이터가 필요함)]
        (아래 엑셀 구조를 참고하여 어떤 테이블과 컬럼이 필요한지 맥락을 파악하십시오. 단, 절대 1번 규칙[스키마 전용 쿼리]을 어기고 엑셀 모양대로 PIVOT하거나 데이터를 변형하지 마십시오.)
        {json.dumps(empty_cells_info, ensure_ascii=False)}

        [사용자 질문/지시]
        {question}

        [출력 지침]
        1. 🚨 **[최우선 판별 규칙 - 지능형 카탈로그 매핑]** 사용자의 [사용자 질문/지시] 속에 포함된 명칭이나 조건어(예: "OOO만 보고", "XXX에서 찾아")를 위 [데이터베이스 스키마 및 힌트]에 제공된 실제 카탈로그 이름들과 대조하여 스스로 검색 모드("search_mode")를 결정하십시오.
           - 질문 속 단어가 **비정형 문서 목록 (ChromaDB에 저장됨)의 파일명이나 타이틀**과 의미상/텍스트상 일치한다면 (예: "2027학년도 모집요강만 보고" -> '2027학년도 모집요강.pdf' 존재 확인), 무조건 "DOC_ONLY"로 강제 설정하십시오.
           - 질문 속 단어가 **MS-SQL 테이블 구조의 테이블명이나 용도**와 의미상/텍스트상 일치한다면 (예: "입학정원만 보고" -> 'ADMISSIONQUOTA' 테이블 존재 확인), 무조건 "DB_ONLY"로 강제 설정하십시오.
           - 위와 같은 명시적 대조가 없거나 두 영역 모두 검색이 필요한 일반적인 경우: "HYBRID"
        2. 만약 "search_mode"가 "DOC_ONLY"라면 "sql_query"는 반드시 "NONE"으로 출력하십시오.
        3. 그 외 정형 통계(지원인원, 등록인원 등)가 필요하다면 엑셀의 빈칸을 모두 채울 수 있는 포괄적인 T-SQL SELECT 문을 작성하십시오.

        오직 순수 JSON 형식으로만 응답해. 마크다운 기호 금지.
        {{
            "search_mode": "DB_ONLY | DOC_ONLY | HYBRID",
            "sql_query": "엑셀의 빈칸을 모두 채울 수 있는 포괄적인 T-SQL SELECT 문 (DOC_ONLY인 경우 'NONE')",
            "target_documents": ["사용자가 특정 문서를 지정한 경우 해당 파일명들"]
        }}
        """
        
        # JSON 응답 강제화를 위한 설정
        generation_config = genai.types.GenerationConfig(
            response_mime_type="application/json",
            response_schema={
                "type": "OBJECT",
                "properties": {
                    "search_mode": {"type": "STRING", "description": "DB_ONLY | DOC_ONLY | HYBRID"},
                    "sql_query": {"type": "STRING", "description": "엑셀의 빈칸을 모두 채울 수 있는 포괄적인 T-SQL SELECT 문 (DOC_ONLY인 경우 'NONE')"},
                    "target_documents": {
                        "type": "ARRAY",
                        "items": {"type": "STRING"},
                        "description": "질문자가 특정 문서를 명시적으로 지정하여 검토를 요구한 경우 일치하는 공식 파일명 리스트. 지정 없으면 빈 배열"
                    }
                },
                "required": ["search_mode", "sql_query", "target_documents"]
            }
        )

        sql_model = genai.GenerativeModel(router_model_name)
            
        no_retry = AsyncRetry(initial=0, maximum=0, multiplier=0, deadline=0)
        sql_response = await sql_model.generate_content_async(
            sql_router_prompt, 
            request_options={"retry": no_retry},
            generation_config=generation_config
        )
        
        router_in_tokens = 0
        router_out_tokens = 0
        if hasattr(sql_response, 'usage_metadata') and sql_response.usage_metadata:
            router_in_tokens = sql_response.usage_metadata.prompt_token_count
            router_out_tokens = sql_response.usage_metadata.candidates_token_count

        # 에러 방어: 변수 사전 초기화 (파싱 실패 시 서버 뻗음 방지)
        search_mode = "HYBRID"
        sql_query = "NONE"
        time_router_end = time.time()

        # JSON 파싱
        try:
            response_text = sql_response.text.strip()
            if response_text.startswith("```json"):
                response_text = response_text.replace("```json", "").replace("```", "").strip()
            elif response_text.startswith("```"):
                response_text = response_text.replace("```", "").strip()
                
            parsed_json = json.loads(response_text)
            search_mode = parsed_json.get("search_mode", "HYBRID")
            sql_query = parsed_json.get("sql_query", "NONE")
            target_documents = parsed_json.get("target_documents", [])
            
            if target_documents:
                print(f"[EXCEL-AUTOFILL] 🎯 타겟 문서 지정됨: {target_documents}")
            
            if search_mode == "DOC_ONLY":
                sql_query = "NONE"
            
            time_router_end = time.time()
            print(f"[EXCEL-AUTOFILL] ⏱️ 1단계(라우터 AI 판단) 소요 시간: {time_router_end - start_time:.2f}초")
            print(f"[EXCEL-AUTOFILL] 🔀 라우터 판별 검색 모드: {search_mode}")
            
            # [2차 절대 방어선] 정규식을 이용해 대괄호 [ ] 내부의 모든 종류의 공백(띄어쓰기, 줄바꿈, 탭, 눈에 보이지 않는 유니코드 공백 등) 강제 삭제
            # 예: [수험 번호] 또는 [ 수험번호] -> [수험번호]
            if sql_query and sql_query != "NONE":
                # [1.5차 절대 방어선] 따옴표('') 내부의 줄바꿈(\n) 강제 제거 (문자열 중간 엔터 에러 방지)
                sql_query = re.sub(r"'[^']*'", lambda m: m.group(0).replace('\n', '').replace('\r', ''), sql_query)

                # \s+ 로 일반 공백 처리 후, 모든 형태의 공백을 리스트로 명시하여 완벽 제거
                sql_query = re.sub(r'\[([^\]]+)\]', lambda m: '[' + re.sub(r'[\s\u200B\u200C\u200D\uFEFF\u3000]+', '', m.group(1)) + ']', sql_query)
                sql_query = sql_query.replace("[ ", "[").replace(" ]", "]")
                
                # [3차 방어선] 스키마 기반 동적 오타 자동 교정 (Self-Healing)
                import difflib
                try:
                    # 1. DB에서 실제 유효한 컬럼명 리스트 확보
                    valid_columns = []
                    if conn:
                        col_cursor = conn.cursor()
                        col_cursor.execute("SELECT column_name FROM Sys_ColumnCatalog")
                        valid_columns = [row[0] for row in col_cursor.fetchall()]
                    
                    if valid_columns:
                        # CTE 등에서 자주 쓰이는 가상 컬럼명/별칭 강제 추가 (오타 교정 대상에서 제외)
                        valid_columns.extend(["상세전형명", "지원인원", "등록인원", "학과명", "모집인원", "전형명", "지원율", "등록율", "합계", "Q", "A", "QUOTA", "APPLICANT", "APPLICANT_BASE", "EXCEL_ADMISSIONQUOTA"])
                        
                        # 2. AS 별칭 부분 보호 (AS [별칭] 부분을 임시 문자열로 치환하여 교정에서 제외)
                        # 공백이 없거나(AS[별칭]) 테이블명.컬럼명([A].[컬럼명]) 등 보호
                        alias_placeholders = {}
                        def alias_replacer(match):
                            placeholder = f"__ALIAS_{len(alias_placeholders)}__"
                            alias_placeholders[placeholder] = match.group(0)
                            return placeholder
                            
                        temp_query = re.sub(r'(?i)\bAS\s*\[([^\]]+)\]', alias_replacer, sql_query)
                        
                        # 3. 남은 쿼리 내의 모든 [컬럼명] 추출 및 유사도 대조
                        def column_healer(match):
                            col_name = match.group(1)
                            # 실제 컬럼 목록에 없으면 가장 유사한 컬럼으로 교정 (정확도 75% 이상)
                            if col_name not in valid_columns:
                                matches = difflib.get_close_matches(col_name, valid_columns, n=1, cutoff=0.75)
                                if matches:
                                    print(f"[동적 교정 완료] AI 오타 감지 및 수정: {col_name} -> {matches[0]}")
                                    return f"[{matches[0]}]"
                            return match.group(0)
                            
                        temp_query = re.sub(r'\[([^\]]+)\]', column_healer, temp_query)
                        
                        # 4. 보호했던 AS 별칭 복원
                        for placeholder, original_alias in alias_placeholders.items():
                            temp_query = temp_query.replace(placeholder, original_alias)
                            
                        sql_query = temp_query
                        
                except Exception as col_err:
                    print(f"[EXCEL-AUTOFILL] 동적 컬럼 교정 실패 (무시됨): {col_err}")

        except Exception as e:
            print(f"[EXCEL-AUTOFILL] SQL JSON 파싱 오류: {e}")
            sql_query = "NONE"
            
        original_sql_query = sql_query
            
        # 8127 구문 에러 원천 차단을 위해 ORDER BY 절 강제 제거
        raw_sql_query = re.sub(r'(?i)\s*ORDER\s+BY\s+.*', '', sql_query).strip()
        if not raw_sql_query.endswith(';'):
            raw_sql_query += ';'
            
        print_sql = format_sql_query(raw_sql_query)
        
        # SSMS 검증을 위한 쿼리 자동 저장 (가독성을 위해 포맷팅 적용)
        try:
            with open('latest_query.sql', 'w', encoding='utf-8') as f:
                f.write(print_sql)
        except Exception as f_err:
            print(f"[EXCEL-AUTOFILL] latest_query.sql 저장 실패: {f_err}")

        print(f"[EXCEL-AUTOFILL] [QUERY] AI 원본 쿼리문:")
        print("--- T-SQL START ---")
        print(format_sql_query(sql_query))
        print("--- T-SQL END ---")
        
        print(f"[EXCEL-AUTOFILL] [QUERY] 파이썬 보정 후 실제 실행 쿼리문:")
        print("--- T-SQL START ---")
        print(print_sql)
        print("--- T-SQL END ---")
        
        # 3. 쿼리 실행
        sql_context = "추출된 DB 데이터 없음"
        if (raw_sql_query.upper().startswith("SELECT") or raw_sql_query.upper().startswith("WITH")) and conn:
            try:
                print("[EXCEL-AUTOFILL] 🔍 3단계: 통계 DB 데이터 획득 중...")
                cursor = conn.cursor()
                cursor.execute(raw_sql_query)
                cols = [column[0] for column in cursor.description]
                rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
                
                for r in rows:
                    for k, v in r.items():
                        if hasattr(v, 'isoformat'):
                            r[k] = v.isoformat()
                        elif isinstance(v, Decimal):
                            r[k] = float(v)
                            
                # 파이썬 레벨에서 DB 결과 강제 정렬 (ORDER BY 8127 에러 회피 및 하드코딩 제거)
                # 쿼리의 SELECT 컬럼 순서(보통 대분류->소분류)를 기준으로 자동 그룹핑 정렬을 수행합니다.
                if rows:
                    rows.sort(key=lambda x: tuple(str(v) for v in x.values()))
                        
                sql_context = json.dumps(rows, ensure_ascii=False)
            except Exception as e:
                # DB 에러 발생 시 2단계 AI로 넘어가면 가짜(환각) 데이터를 생성하므로 즉시 에러를 던져 중단합니다.
                error_msg = f"DB 쿼리 실행 중 오류가 발생했습니다: {str(e)}\n\n[AI 원본 쿼리]\n{original_sql_query}\n\n[실제 실행 쿼리]\n{sql_query}"
                print(f"[EXCEL-AUTOFILL] DB 에러 (즉시 중단): {e}")
                raise HTTPException(status_code=500, detail=error_msg)

        # [NEW] 3.5단계: 비정형 DB(ChromaDB)에서 문서 검색
        print("[EXCEL-AUTOFILL] 📚 3.5단계: 비정형 DB 문서 검색 중...")
        chroma_context = "검색된 비정형 문서가 없습니다."
        # 파일명에서 학년도 자동 추출, 없으면 질문에서 추출
        year_match = re.search(r'(202\d)', file.filename)
        if not year_match:
            year_match = re.search(r'(202\d)', question)
        target_year = year_match.group(1) if year_match else "ALL"

        chroma_context = ""
        try:
            if search_mode == "DB_ONLY":
                chroma_context = "사용자 지시로 인해 비정형 문서 검색이 생략되었습니다."
                print("[EXCEL-AUTOFILL] 🚫 사용자 지시(DB_ONLY)에 따라 비정형 DB 검색 생략.")
            else:
                from main import rule_db, reference_db
                
                # 파일명 한글 깨짐(Mojibake) 버그를 우회하기 위해 filename 필터 전면 삭제.
                # 오직 의미론적 유사도(Semantic Search)로만 문서 탐색
                # 모집인원 하드코딩 제거: 사용자의 질문 전체를 쿼리로 사용하여 범용적인 검색 수행
                query_texts = [question]
                
                if target_year != "ALL":
                    where_clause = {"year": {"$in": [target_year, "ALL"]}}
                    res_rule = rule_db.query(query_texts=query_texts, n_results=100, where=where_clause)
                    res_ref = reference_db.query(query_texts=query_texts, n_results=100, where=where_clause)
                else:
                    res_rule = rule_db.query(query_texts=query_texts, n_results=100)
                    res_ref = reference_db.query(query_texts=query_texts, n_results=100)
                
                doc_texts = []
                for res in [res_rule, res_ref]:
                    if res and res.get("documents") and res["documents"][0]:
                        for i, doc_text in enumerate(res["documents"][0]):
                            meta = res["metadatas"][0][i] if res.get("metadatas") else {}
                            fname = meta.get("filename", "이름없음")
                            
                            # [NEW] 타겟 문서 지정 (Target Document Enforcement)
                            if target_documents:
                                fname_lower = fname.lower()
                                target_clean = [re.sub(r'[^a-zA-Z0-9가-힣]', '', t).lower() for t in target_documents]
                                fname_only_clean = re.sub(r'[^a-zA-Z0-9가-힣]', '', fname.replace(".pdf", "")).lower()
                                if not any((tc in fname_only_clean or tc in fname_lower) for tc in target_clean):
                                    continue
                                    
                            doc_texts.append(f"[{fname}]\n{doc_text}")
                    
                if doc_texts:
                    # 100개 이상의 청크(전체 PDF 분량)를 그대로 결합하여 테이블 누락 방지!
                    chroma_context = "\n\n".join(doc_texts[:200])
                    print(f"[EXCEL-AUTOFILL] 📚 검색 성공: {len(doc_texts)}개의 문서 문단(전체) 참조")
                else:
                    print("[EXCEL-AUTOFILL] 📚 검색된 문서 문단이 없습니다.")
        except Exception as e:
            print(f"[EXCEL-AUTOFILL] ⚠️ ChromaDB 문서 검색 오류 (무시됨): {e}")

        time_db_end = time.time()
        print(f"[EXCEL-AUTOFILL] ⏱️ 2단계(DB 및 문서 검색) 소요 시간: {time_db_end - time_router_end:.2f}초")

        # 4. 최종 매핑 AI 구동
        print(f"[EXCEL-AUTOFILL] 🧠 4단계: 최종 엑셀 매핑 및 답변 생성 중... (사용 모델: {model_name})")

        final_prompt = f"""
        너는 대동대학교 입시처의 [엑셀 데이터 매핑 AI]야.

        [사용자 질문]
        {question}

        [관리자 등록 최상위 사전지식]
        {pre_supplemental_text if pre_supplemental_text else "등록된 사전지식 없음"}
        - 위 사전지식은 엑셀에 데이터를 매핑할 때 반드시 준수해야 하는 행정 룰입니다.

        [🚨 사용자 특별 지시사항 (최우선 적용 규칙) 🚨]
        - 사용자의 [사용자 질문] 내에 엑셀 값 입력과 관련된 명시적인 조건이나 제약(예: "제한없음은 제한없음으로 써라", "하이픈(-)은 제외하라" 등)이 있다면, **아래의 모든 기본 출력 지침보다 사용자의 이 지시를 최우선으로, 그리고 완벽하게 따르십시오!**
        - 단, 사용자의 질문이 엑셀 값에 대한 명시적 조건이 아니라 단순한 대화형 질문(예: "왜 이렇게 많아?", "이거 줄바꿈이니?" 등)일 경우, 사용자의 질문 내용 때문에 엑셀 매핑을 포기하거나 숫자를 임의로 0으로 만들어서는 안 됩니다. 무조건 DB 데이터를 있는 그대로 매핑하십시오.

        [주의 사항 (환각 방지 및 병합된 셀 처리)]
        - 사용자가 데이터 추출을 요청한 대상 연도는 '{target_year}'학년도입니다. 
        - 💡 **[데이터 매핑 최우선 및 결측치 규칙]**: 빈 셀을 채울 때는 무조건 [1. 정형 DB 추출 통계 데이터]의 값을 최우선적으로 매핑하십시오. DB 데이터에 값(예: 123)이 명확히 존재하는데도 사용자 질문 내용에 휘둘리거나 엑셀의 헤더명과 완벽하게 일치하지 않는다는 이유로 임의로 0으로 무시하고 덮어쓰는 행위는 절대 금지됩니다! (유연하게 의미를 파악하여 매핑하세요.) 오직 **DB와 PDF 양쪽 모두에 데이터가 존재하지 않을 때만** 다른 연도 데이터를 지어내지 말고 0으로 처리하십시오.
        - 💡 **[PDF 표 병합(Merge) 데이터 추론 규칙]**: 비정형 PDF 문서의 표(Table) 특성상, 인접한 두 학과의 데이터가 동일할 경우 셀이 병합되어 텍스트 추출 시 **첫 번째 학과에만 값이 적히고 그 아래 학과는 값이 누락된 것(빈칸)처럼 보일 수 있습니다.** 만약 특정 학과 전형의 값이 비어있다면, **반드시 바로 위(또는 근처) 학과의 동일 전형 값을 확인하고, 병합된 셀이라고 판단되면 그 값을 똑같이 상속(복사)하여 채워 넣으십시오.**

        [데이터베이스 스키마 및 힌트]
        (각 테이블과 컬럼의 의미(ai_description)를 참고하여 아래 추출된 통계 데이터의 맥락을 이해하세요.)
        {dynamic_schema}

        [1. 정형 DB 추출 통계 데이터]
        {sql_context}
        [2. 비정형 PDF 추출 규정 문서 참조 내용]
        (🎯 핵심 지시: DB 통계와 함께 아래 PDF 문서 텍스트를 교차 검증하여 엑셀의 빈칸을 채우십시오.)
        {chroma_context}
        
        [첨부 엑셀 파일명]
        {file.filename}

        [현재 엑셀 파일에 존재하는 데이터(헤더 및 기존 값)]
        {existing_data_context if existing_data_context else "없음"}

        [엑셀 빈 셀 구조]
        (아래 빈 셀 좌표가 주어졌다면 해당 좌표에 맞게 채우고, 만약 비어있다면([]) 위 헤더 구조를 보고 AI가 스스로 적절한 행(예: 2행, 3행...)을 새롭게 무한대로 추가해가며 표를 완성해서 좌표를 생성하세요.)
        {json.dumps(empty_cells_info, ensure_ascii=False)}
        
        [출력 지침]
        
        출력은 반드시 안내 텍스트로 시작하되, 
        안내 텍스트는 **무조건 3줄 이내로 극도로 짧게 핵심만 요약**해. (절대 학과별로 길게 나열하지 마십시오! 최대 출력 한도에 걸립니다.)
        
        그 다음, 엑셀에 주입할 실제 데이터는 **반드시 마지막에 순수한 JSON 배열 형식으로만** 작성해.
        - **[중요: 의미 보존형 Key 네이밍 규칙 (Self-Correction)]** 출력 길이를 압축하면서도, 100줄 이상 장문 출력 시 AI 스스로 열(Column)이 밀리는 환각(Shift)을 원천 방지하기 위해 **각 행의 열 Key는 반드시 `[알파벳]_[컬럼명]` 형태로 결합**하여 작성하십시오. 절대 의미 없는 알파벳 단독(`"A"`)으로 쓰지 마십시오! 스스로 의미를 각인해야 합니다.
        - **[표기 일관성 범용 강제]** 데이터 표기 방식(예: 연도, 부서명 포맷 등)은 최초 1행에서 결정한 형식을 마지막 행까지 100% 동일하게 유지(Consistency)하십시오. 중간에 표기법을 임의로 변경하지 마십시오.
        - **[🔥 강력한 의미 유추 매핑 규칙 🔥]** DB 통계 결과의 전형명(예: `만학도`, `장애인대상자`)과 엑셀 헤더명(예: `만학도(7호)`, `장애인대상자(4호)`)이 글자 토씨 하나까지 완벽히 일치하지 않더라도, 괄호 안의 부가설명이나 기호 차이일 뿐 **핵심 의미가 같다면 반드시 동일한 항목으로 간주하여 데이터를 매핑하십시오.** 절대로 "이름이 다르다"며 0으로 누락시켜서는 안 됩니다! 데이터를 최대한 유연하고 똑똑하게 추론하여 실제 추출된 숫자 데이터를 적극적으로 채우십시오.
        - **[🚨 데이터 행(Row) 완전성 보장 (세로 생략 금지) 🚨]** 빈 엑셀 템플릿에 행을 창조해 낼 때, 문서에 존재하는 모든 학과와 전형 조합을 **단 하나도 누락시키지 말고 끝까지(100%) 추출하십시오.** 출력이 150줄을 넘어가더라도 중간에 힘들다고 임의로 생략하거나 멈추는 행위(Lazy Generation)는 절대 금지됩니다! 무조건 100% 전부 배열에 담으십시오.
        - **[🚨 데이터 열(Column) 완전성 보장 (가로 생략 금지) 🚨]** 각 행(row)을 출력할 때, [엑셀 빈 셀 구조]에 제시된 **마지막 열(예: `BI_`)까지 모든 컬럼(열) Key를 단 하나도 빠짐없이 100% 전부 적어 넣으십시오.** 뒤쪽에 위치한 '서해5도', '다자녀가정', '합계', '지원율' 등의 컬럼들도 값이 0이거나 비어있더라도 절대 임의로 생략(Truncation)하지 마십시오. 컬럼이 누락되면 시스템이 파괴됩니다.
        - **[🔥 치명적 규칙: "row" 키 절대 필수 🔥]** 답변 배열의 각 JSON 객체 안에는 무조건 `"row": [해당 행 번호]` 키를 **제일 첫 번째 항목**으로 반드시 포함하십시오. (예: `{{"row": 5, ...}}`) 만약 `"row"` 번호가 단 하나라도 누락되면 엑셀에 데이터를 매핑할 수 없어서 파멸적인 오류가 발생합니다! 빈 좌표를 참고하거나, 새 행을 창조할 때는 이전 행에서 `+1`씩 증가시켜서라도 **무조건 모든 객체에 "row" 키를 집어넣으십시오.**
        - 기본 규칙 (단, 위 '사용자 특별 지시사항'과 충돌 시 사용자 지시를 100% 우선함): 데이터 값이 숫자일 경우 가급적 정수형(`int`)으로 입력.
        - 💡 **[비정형 문서 기호/단어 의미 규칙]**: 비정형 문서 표에서 하이픈(`-`) 기호는 '데이터 값이 존재하지 않음(해당 전형으로 모집하지 않음)'을 의미하므로, 특별한 사용자 지시가 없다면 숫자 `0`과 동일하게 치환하여 취급할 수 있습니다. 또한 `제한없음`이라는 단어는 모집 인원이 무제한이어서 값이 숫자로 고정되어 있지 않음을 의미합니다.
        - 💡 **[데이터 충돌 시 우선순위 규칙]**: 만약 [정형 DB 추출 통계 데이터]와 [비정형 PDF 추출 규정 문서]의 내용이 서로 상이하거나 충돌할 경우, 최신 통계 데이터인 **[정형 DB 추출 통계 데이터]의 값을 최우선으로 참고**하십시오.
        - **[필수: 문자열 보존 규칙] 만약 사용자가 "제한없음" 문자열이나 "하이픈(-)", 빈칸을 유지하라고 지시했다면, 강제로 숫자로 변환하지 말고 사용자의 요구대로 문자열을 그대로 매핑하거나 해당 셀을 과감히 생략하십시오.**
        - 만약 사용자가 "해당되지 않으면 0으로 채워줘"라고 지시했다면, 해당 영역의 빈칸이나 하이픈(-)에 0 값을 적극적으로 출력하십시오.
        - **[🚨 병합 셀(Merge) 상속 및 예외 규칙 🚨]** 
          1) 표에서 값이 비어있는 칸을 발견하면 무조건 '같은 전형(Column)의 윗줄'을 확인하십시오. 윗줄이 '제한없음' 값을 가진 병합된 셀이라면 0으로 채우지 말고 똑같이 복사(상속)하십시오.
          2) **[중요 예외 1 - 모집시기 간 억지 상속 절대 금지]** 윗줄의 값을 상속받는 것은 오직 **동일한 모집시기(예: 같은 '수시1차' 내부)**에서만 가능합니다! '수시2차' 표의 '제한없음' 값을 '정시' 표 빈칸으로 끌어와서 억지로 채워넣는 오류를 절대 범하지 마십시오. (특히 '정시'는 정원외 전형을 아예 모집하지 않는 경우가 많으므로 주의하세요.)
          3) **[중요 예외 2 - 없는 전형 상속 금지]** 해당 전형 자체가 아예 모집을 하지 않아서 원본 문서에 하이픈(-)으로 도배되어 있거나 완전히 통째로 비어있는 경우에는 절대로 억지로 상속하지 마십시오. 명확히 모집하지 않는 전형은 과감히 생략하거나 하이픈을 유지해야 합니다.
        ```json
        [
          {{"row": 2, "A_입시학년도": 2026, "C_지원학과명": "간호학부", "D_정원내외구분": "정원내", "E_전형명": "일반고"}},
          {{"row": 3, "A_입시학년도": 2026, "C_지원학과명": "응급구조과", "D_정원내외구분": "정원내", "E_전형명": "일반고"}}
        ]
        ```
        """
        
        final_model = genai.GenerativeModel(model_name)
        no_retry = AsyncRetry(initial=0, maximum=0, multiplier=0, deadline=0)
        final_response = await final_model.generate_content_async(
            final_prompt,
            request_options={"timeout": 600, "retry": no_retry}
        )
        final_text = final_response.text
        
        final_in_tokens = 0
        final_out_tokens = 0
        if hasattr(final_response, 'usage_metadata') and final_response.usage_metadata:
            final_in_tokens = final_response.usage_metadata.prompt_token_count
            final_out_tokens = final_response.usage_metadata.candidates_token_count
        
        # 5. 매핑 추출 및 엑셀 주입
        mapping_result = {}
        answer_text = final_text
        
        # 새로운 JSON 포맷 파싱 시도 (행 단위 2D 압축 포맷)
        json_match = re.search(r'```(?:json)?\s*(\[.*?\])\s*```', final_text, re.DOTALL)
        if not json_match:
            # 백틱이 없거나 닫히지 않았을 수 있음
            json_match = re.search(r'(\[\s*\{\s*"row".*)', final_text, re.DOTALL)

        if json_match:
            answer_text = final_text.replace(json_match.group(0), "").strip()
            try:
                data_list = json.loads(json_match.group(1))
                for item in data_list:
                    row_idx = item.get("row")
                    if row_idx is None:
                        continue
                    for col_key, val in item.items():
                        if col_key != "row" and isinstance(col_key, str):
                            # 의미 보존형 Key("A_입시학년도" 또는 "A")에서 앞의 알파벳만 추출
                            match = re.match(r'^([A-Z]+)', col_key.upper())
                            if match:
                                coord = f"{match.group(1)}{row_idx}"
                                mapping_result[coord] = val
            except Exception as e:
                print(f"[EXCEL-AUTOFILL] 신규 JSON 2D 파싱 실패(정규식 폴백 시도): {e}")

        # JSON 파싱 실패했거나(Truncated), 파싱했지만 비어있는 경우 정규식으로 직접 추출 (Truncated 방어)
        if not mapping_result:
            print("[EXCEL-AUTOFILL] ⚠️ 강력한 정규식 추출 모드 가동 (Truncated JSON 방어)")
            row_blocks = re.findall(r'\{\s*"row"\s*:\s*(\d+)(.*?)\}', final_text, re.DOTALL)
            for row_idx_str, col_data in row_blocks:
                kv_pattern = r'"([a-zA-Z]+)[^"]*"\s*:\s*(?:"([^"]+)"|([0-9\.-]+))'
                for match in re.finditer(kv_pattern, col_data):
                    col_letter = match.group(1).upper()
                    str_val = match.group(2)
                    num_val = match.group(3)
                    
                    val = str_val if str_val is not None else num_val
                    if val is not None:
                        if isinstance(val, str) and re.match(r'^-?\d+(\.\d+)?$', val):
                            val = float(val) if '.' in val else int(val)
                        mapping_result[f"{col_letter}{row_idx_str}"] = val

        # 기존 방식 fallback
        if not mapping_result:
            map_match = re.search(r'\[EXCEL_MAP_START\](.*?)\[EXCEL_MAP_END\]', final_text, re.DOTALL)
            if map_match:
                json_str = map_match.group(1).strip()
                answer_text = final_text.replace(map_match.group(0), "").strip()
                pattern = r'"([a-zA-Z]+[0-9]+)"\s*:\s*([0-9\.-]+)'
                matches = re.findall(pattern, json_str)
                for k, v in matches:
                    mapping_result[k.upper()] = float(v) if '.' in v else int(v)

        filled_count = 0
        for coord, val in mapping_result.items():
            try:
                ws[coord].value = val
                filled_count += 1
            except Exception as e:
                print(f"[EXCEL-AUTOFILL] 셀 {coord} 쓰기 실패: {e}")
                
        def calc_cost(model, in_tok, out_tok):
            if "3.5" in model:
                return (in_tok / 1000000 * 1.25 + out_tok / 1000000 * 5.0) * 1400
            else:
                return (in_tok / 1000000 * 0.075 + out_tok / 1000000 * 0.3) * 1400
        
        router_cost = calc_cost(router_model_name, router_in_tokens, router_out_tokens)
        final_cost = calc_cost(model_name, final_in_tokens, final_out_tokens)
        total_cost = router_cost + final_cost
        total_tokens = router_in_tokens + router_out_tokens + final_in_tokens + final_out_tokens

        end_time = time.time()
        latency_ms = int((end_time - start_time) * 1000)
        print(f"[EXCEL-AUTOFILL] ⏱️ 엑셀 채우기 총 소요 시간: {end_time - start_time:.2f}초")
        print(f"[EXCEL-AUTOFILL] 🎉 하이브리드 엑셀 채우기 성공! ({latency_ms}ms)")
        
        print("\n==================================================")
        print("🧾 [💸 엑셀 AI 요금 결제 내역 (영수증)]")
        print(f"[라우터 AI] 사용 모델: {router_model_name}")
        print(f"  - 입력: {router_in_tokens:,} / 출력: {router_out_tokens:,} (예상 비용: 약 {router_cost:.3f} 원)")
        print("--------------------------------------------------")
        print(f"[최종 매핑 AI] 사용 모델: {model_name}")
        print(f"  - 입력: {final_in_tokens:,} Tokens")
        print(f"  - 출력: {final_out_tokens:,} Tokens")
        print(f"  - 예상 비용: 약 {final_cost:.3f} 원")
        print("--------------------------------------------------")
        print(f"🔹 총 소모 토큰 합계: {total_tokens:,} Tokens")
        print(f"💰 총 예상 청구 비용: 약 {total_cost:.3f} 원 (KRW)")
        print("==================================================\n")

        if conn:
            try:
                def safe_str(s):
                    return s.encode('cp949', errors='replace').decode('cp949') if s else ""
                
                log_cursor = conn.cursor()
                log_cursor.execute('''
                    INSERT INTO Sys_AIAuditLog (
                        created_at, user_role, question, answer, scraped_context, sql_query,
                        latency_ms, model_name, router_model_name,
                        router_in_tokens, router_out_tokens,
                        chat_in_tokens, chat_out_tokens, total_tokens,
                        estimated_cost
                    ) VALUES (
                        GETDATE(), ?, ?, ?, ?, ?,
                        ?, ?, ?,
                        ?, ?,
                        ?, ?, ?,
                        ?
                    )
                ''', (
                    safe_str(user_role), safe_str(question), safe_str(answer_text[:2000]), safe_str(f"Mode: {search_mode}"), safe_str(original_sql_query),
                    latency_ms, safe_str(model_name), safe_str(router_model_name),
                    router_in_tokens, router_out_tokens,
                    final_in_tokens, final_out_tokens, total_tokens,
                    total_cost
                ))
                conn.commit()
            except Exception as e:
                print(f"[EXCEL-AUTOFILL] Audit Log 저장 실패: {e}")

        if filled_count > 0:
            print(f"[EXCEL-AUTOFILL] ✅ 5단계: AI 매핑 완료 (총 {filled_count}개 셀 주입 성공)")
        else:
            print("[EXCEL-AUTOFILL] ⚠️ 매핑된 데이터가 없습니다.")
            
        # [디버깅용] 로컬 temp 폴더에 원본 파일 강제 저장
        try:
            temp_dir = os.path.join(os.getcwd(), "temp")
            os.makedirs(temp_dir, exist_ok=True)
            debug_file_path = os.path.join(temp_dir, f"DEBUG_Autofill_{file.filename}")
            wb.save(debug_file_path)
            print(f"[EXCEL-AUTOFILL] 💾 서버 저장 원본 엑셀 경로: {debug_file_path}")
        except Exception as e:
            print(f"[EXCEL-AUTOFILL] ⚠️ 임시 파일 저장 실패: {e}")
            
        # 6. Base64 인코딩 및 반환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_base64 = base64.b64encode(output.read()).decode('utf-8')
        
        latency_ms = int((time.time() - start_time) * 1000)
        time_final_end = time.time()
        print(f"[EXCEL-AUTOFILL] ⏱️ 3단계(최종 엑셀 매핑 AI) 소요 시간: {time_final_end - time_db_end:.2f}초")
        print(f"[EXCEL-AUTOFILL] ⏱️ 엑셀 채우기 총 소요 시간: {time_final_end - start_time:.2f}초")
        print(f"[EXCEL-AUTOFILL] 🎉 하이브리드 엑셀 채우기 성공! ({latency_ms}ms)")
        print("[EXCEL-AUTOFILL] ==================================================\n")
        
        return {
            "status": "success",
            "answer": answer_text,
            "log_id": -1, # 로그 저장은 생략하거나 필요 시 추가
            "latency_ms": latency_ms,
            "excel_base64": excel_base64,
            "excel_filename": f"AutoFilled_{file.filename}"
        }
        
    except asyncio.CancelledError:
        print(f"[EXCEL-AUTOFILL] 🛑 사용자에 의해 요청이 취소되었습니다.")
        raise
    except Exception as e:
        print(f"[EXCEL-AUTOFILL] ❌ 에러 발생: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            try:
                conn.close()
            except:
                pass
